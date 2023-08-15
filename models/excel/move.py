import openpyxl

def copy_columns(source_path, dest_path, num_columns):
    # Load the source workbook
    source_workbook = openpyxl.load_workbook(source_path)
    source_sheet = source_workbook.active

    # Create a new workbook for the destination
    dest_workbook = openpyxl.Workbook()
    dest_sheet = dest_workbook.active

    # Copy the first n columns
    for row in source_sheet.iter_rows(max_col=num_columns, values_only=True):
        dest_sheet.append(row)

    # Save the destination workbook
    dest_workbook.save(dest_path)

    print(f"{num_columns} columns copied from '{source_path}' to '{dest_path}'")

if __name__ == "__main__":
    source_file = "faculty.xlsx"  # Replace with your source file path
    dest_file = "faculty_new.xlsx"  # Replace with your destination file path
    num_columns_to_copy = 6  # Replace with the number of columns you want to copy

    copy_columns(source_file, dest_file, num_columns_to_copy)

